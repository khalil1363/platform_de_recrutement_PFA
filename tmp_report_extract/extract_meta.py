# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook
import pdfplumber

out = Path(r"c:\Users\lpatron\OneDrive\Desktop\mohamed\daam\tmp_report_extract")

p = Path(r"c:\Users\lpatron\OneDrive\Desktop\Rapport du Profil Pro 2 de Achref JALLELI.pdf")
with pdfplumber.open(p) as pdf:
    chunks = []
    for i in [5, 6, 7, 8, 9, 10, 11]:
        if i < len(pdf.pages):
            chunks.append(f"===== PAGE {i+1} =====\n" + (pdf.pages[i].extract_text() or ""))
    (out / "profil_mid.txt").write_text("\n".join(chunks), encoding="utf-8")

wb = load_workbook(
    r"c:\Users\lpatron\OneDrive\Desktop\2026_Suivi Test PSY CC_1.0 (1).xlsx", data_only=True
)
ws = wb["KPIs Commercial Terrain"]
comps = []
for r in range(5, 25):
    comp = ws.cell(r, 2).value
    expected = ws.cell(r, 4).value
    comment = ws.cell(r, 5).value
    comps.append(f"{comp}|{expected}|{comment}")
(out / "competencies_template.txt").write_text("\n".join(comps), encoding="utf-8")
print("comps", len(comps))
for c in comps:
    print(c)
